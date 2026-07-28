import contextlib
import re
import warnings
from datetime import datetime

import openpyxl
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from django.db import transaction
from django.db.models import Count
from django.utils import timezone

from leads.models import Lead, LeadEvent

User = get_user_model()

FAKE_MANAGERS = {"sms junatildi o'chirilgan", "sms junatildi uchirilgan"}

SKIP_SHEETS = {"users"}

NO_PHONE = "Noma'lum"

# Excel "Varonka Etapi" -> (status, sub_status). Kalitlar norm() bilan solishtiriladi.
STATUS_MAP = {
    "murojaat qildi": ("yangi_murojaat", "murojaat_qildi"),
    "ko'rdi / eshitdi": ("yangi_murojaat", "kordi_eshitdi"),
    "ko'rdi/ eshitdi": ("yangi_murojaat", "kordi_eshitdi"),
    "ko'rdi/eshitdi": ("yangi_murojaat", "kordi_eshitdi"),
    "atkaz (ko'rdi/eshitdi)": ("bekor_qilingan", "atkaz_qildi"),
    "atkaz (keldi)": ("bekor_qilingan", "atkaz_qildi"),
    "atkaz qildi": ("bekor_qilingan", "atkaz_qildi"),
    "telifon mavjud emas": ("bekor_qilingan", "atkaz_qildi"),
    "telefon mavjud emas": ("bekor_qilingan", "atkaz_qildi"),
    "uchrashuv belgilandi": ("uchrashuv", "uchrashuv_belgilandi"),
    "keldi": ("uchrashuv", "keldi"),
    "shartnoma qildi": ("jarayon", "shartnoma_qildi"),
    "band qildi": ("jarayon", "band_qildi"),
    "notarius": ("jarayon", "notarius"),
    "uy oldi": ("muvaffaqiyatli", "uy_oldi"),
}

SOURCE_MAP = {
    "instagram": "Instagram",
    "telegram": "Telegram",
    "reklama": "Reklama",
    "suhrob_reklama": "Reklama",
    "tavsiya": "Tavsiya",
    "o'zi kelgan": "Boshqa",
    "noma'lum": "Boshqa",
}

# Ichki kalit -> sarlavha boshlanishi (tartib muhim: aniqrog'i oldinda)
COLUMN_ALIASES = {
    "full_name": ["i.f.sh", "f.i.sh yoki tashkilot", "deriktor f.i.sh", "f.i.sh"],
    "org_name": ["tashkilot nomi", "mahallasi"],
    "phone": ["telifon raqami", "telefon raqami", "telifon", "telefon"],
    "source": ["qayerdan keldi"],
    "subsidiya": ["subsidiya"],
    "status": ["varonka etapi"],
    "manager": ["sotuv menejeri"],
    "meeting": ["uchrashuv belgilandi"],
    "note": ["umumiy ma'lumot"],
    "contact_date": ["gaplashilgan sana", "gaplashilgan san"],
}

# Excel ustuvor rejimida yangilanadigan Lead maydonlari
UPDATABLE_FIELDS = (
    "full_name",
    "source",
    "subsidiya",
    "status",
    "sub_status",
    "owner",
    "assignee",
    "note",
    "meeting_at",
    "contacted_at",
)

COMMENT_RE = re.compile(
    r"\[(\d{2}\.\d{2}\.\d{4});\s*(\d{2}:\d{2})\s*\|\s*([^\]]+)\]\s*💬\s*(.+?)(?=\n\[|\Z)",
    re.DOTALL,
)

# "{Shartnoma qildi} qolgan izoh" ko'rinishidagi status
BRACED_STATUS_RE = re.compile(r"^\{([^}]+)\}\s*(.*)$", re.DOTALL)


def norm(val):
    """Solishtirish uchun: apostroflarni birlashtirish, registr, ortiqcha bo'shliqlar."""
    if val is None:
        return ""
    s = str(val).strip()
    for ch in ("‘", "’", "ʼ", "ʻ", "￼", "`", "´"):
        s = s.replace(ch, "'")
    return re.sub(r"\s+", " ", s).lower()


def normalize(val):
    """Ko'rsatish uchun: apostroflarni birlashtiradi, registrni saqlaydi."""
    if not val:
        return ""
    s = str(val).strip()
    for ch in ("‘", "’", "ʼ", "ʻ", "￼", "`", "´"):
        s = s.replace(ch, "'")
    return s


def clean_phone(val):
    """Excel raqamli katakni to'g'ri o'qiydi: 972667470.0 -> "972667470"."""
    if val is None or val == "":
        return NO_PHONE
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    phone = str(val).split("\n")[0].strip()
    phone = re.sub(r"[^\d+]", "", phone)
    return phone[:30] if phone else NO_PHONE


def legacy_phone(val):
    """Eski (xato) formatlash: float ".0" tufayli oxiriga ortiqcha nol qo'shardi.

    Prodda shu ko'rinishda saqlangan leadlarni topib, raqamini to'g'rilash uchun kerak.
    """
    if not val:
        return NO_PHONE
    phone = str(val).split("\n")[0].strip()
    phone = re.sub(r"[^\d+]", "", phone)
    return phone[:30] if phone else NO_PHONE


def parse_subsidiya(val):
    return norm(val) in ("bor", "jarayonda") if val else False


def make_aware(dt):
    if isinstance(dt, datetime):
        return timezone.make_aware(dt) if timezone.is_naive(dt) else dt
    if isinstance(dt, str):
        dt = dt.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return timezone.make_aware(datetime.strptime(dt, fmt))
            except ValueError:
                continue
    return None


def parse_comments(text):
    if not text:
        return [], ""
    matches = COMMENT_RE.findall(str(text))
    remaining = COMMENT_RE.sub("", str(text)).strip()
    return matches, remaining


def parse_status(raw):
    """(status, sub_status, qoldiq_izoh) qaytaradi. Noma'lum status -> yangi_murojaat."""
    text = normalize(raw)
    extra = ""
    braced = BRACED_STATUS_RE.match(text)
    if braced:
        text, extra = braced.group(1).strip(), braced.group(2).strip()

    mapped = STATUS_MAP.get(norm(text))
    if mapped:
        return mapped[0], mapped[1], extra
    return "yangi_murojaat", "murojaat_qildi", (f"{text} {extra}".strip() if text else extra)


class DryRunRollback(Exception):
    pass


class Command(BaseCommand):
    help = "lid.xlsx dan Lead va LeadEvent larni import qiladi (mavjudlarini telefon bo'yicha yangilaydi)"

    def add_arguments(self, parser):
        parser.add_argument("--file", default="lid.xlsx", help="Excel fayl yo'li (default: lid.xlsx)")
        parser.add_argument("--dry-run", action="store_true", help="Bazaga yozmasdan tekshirish (oxirida rollback)")
        parser.add_argument("--reset", action="store_true", help="Import oldidan barcha leadlarni o'chiradi")
        parser.add_argument(
            "--sheet",
            action="append",
            default=None,
            help="Faqat shu varaq(lar) (bir necha marta berish mumkin). Default: 'Users' dan boshqa hammasi",
        )
        parser.add_argument(
            "--fill-empty-only",
            action="store_true",
            help="Mavjud leadda faqat bo'sh maydonlar to'ldirilsin (default: excel ustuvor, farqlar qayta yoziladi)",
        )
        parser.add_argument("--no-events", action="store_true", help="LeadEvent yozilmasin")

    def handle(self, *args, **options):
        warnings.filterwarnings("ignore")
        for stream in (self.stdout, self.stderr):
            if hasattr(stream, "reconfigure"):
                with contextlib.suppress(Exception):
                    stream.reconfigure(encoding="utf-8")

        self.dry_run = options["dry_run"]
        self.excel_wins = not options["fill_empty_only"]
        self.write_events = not options["no_events"]

        wb = openpyxl.load_workbook(options["file"], data_only=True)

        if options["sheet"]:
            wanted = {norm(s) for s in options["sheet"]}
            sheets = [ws for ws in wb.worksheets if norm(ws.title) in wanted]
            missing = wanted - {norm(ws.title) for ws in sheets}
            if missing:
                self.stderr.write(self.style.ERROR(f"Varaq topilmadi: {missing}. Mavjud: {wb.sheetnames}"))
                return
        else:
            sheets = [ws for ws in wb.worksheets if norm(ws.title) not in SKIP_SHEETS]

        self.users = {norm(u.full_name): u for u in User.objects.all() if u.full_name}
        self.unmatched = {}
        self.sheet_rows = {}
        self.stats = dict.fromkeys(
            ("created", "updated", "unchanged", "skipped", "duplicates", "phones_fixed", "errors", "events"), 0
        )

        if self.dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN: baza o'zgartirilmaydi, oxirida rollback qilinadi"))
        if self.excel_wins:
            self.stdout.write(
                self.style.WARNING("EXCEL USTUVOR: farq qilgan maydonlar exceldagisiga yoziladi (bo'sh katak tegmaydi)")
            )

        try:
            with transaction.atomic():
                if options["reset"]:
                    deleted, _ = Lead.objects.all().delete()
                    self.stdout.write(self.style.WARNING(f"{deleted} ta yozuv o'chirildi"))

                entries = self._collect(sheets)
                for entry in entries:
                    try:
                        self._apply_row(entry)
                    except Exception as exc:
                        self.stats["errors"] += 1
                        self.stderr.write(f"[{entry['sheet']}:{entry['row_num']}] xato: {exc}")

                self._report_by_sheet(entries)

                if self.dry_run:
                    raise DryRunRollback
        except DryRunRollback:
            pass

        s = self.stats
        self.stdout.write(
            self.style.SUCCESS(
                f"\nJami: {s['created']} yaratildi | {s['updated']} yangilandi | {s['unchanged']} o'zgarishsiz | "
                f"{s['skipped']} bo'sh qator | {s['duplicates']} takroriy telefon | "
                f"{s['phones_fixed']} telefon to'g'rilandi | "
                f"{s['errors']} xato | {s['events']} event"
            )
        )
        self._report_duplicate_leads()

        if self.unmatched:
            total = sum(self.unmatched.values())
            self.stdout.write(
                self.style.WARNING(
                    f"\n{len(self.unmatched)} ta menejer DB da topilmadi (jami {total} ta lead owner=None):"
                )
            )
            for name, cnt in sorted(self.unmatched.items(), key=lambda x: -x[1]):
                self.stdout.write(f'   "{name}" — {cnt} ta')
        if self.dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN yakunlandi — baza o'zgarmadi."))

    # --- varaq ---------------------------------------------------------

    def _build_colmap(self, header):
        headers = [(i, norm(h)) for i, h in enumerate(header) if h and norm(h)]
        colmap, taken = {}, set()
        for key, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                match = next((i for i, h in headers if i not in taken and h.startswith(alias)), None)
                if match is not None:
                    colmap[key] = match
                    taken.add(match)
                    break
        return colmap

    def _collect(self, sheets):
        """Barcha varaqlarni o'qib, kalit (telefon yoki ism) bo'yicha dedup qiladi — oxirgi qator g'olib.

        Excelning o'zida bir telefon bir necha marta uchraydi; dedupsiz ular har importda
        navbat bilan bir-birini qayta yozadi va log doim "yangilandi" ko'rsatadi.
        """
        by_key = {}
        for ws in sheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            colmap = self._build_colmap(rows[0])
            if "full_name" not in colmap and "org_name" not in colmap:
                self.stdout.write(self.style.ERROR(f"[{ws.title}] ism ustuni topilmadi, o'tkazildi: {rows[0]}"))
                continue

            self.sheet_rows[ws.title] = len(rows) - 1
            for row_num, row in enumerate(rows[1:], start=2):
                data = {key: (row[i] if i < len(row) else None) for key, i in colmap.items()}
                try:
                    entry = self._parse_row(data, ws.title, row_num)
                except Exception as exc:
                    self.stats["errors"] += 1
                    self.stderr.write(f"[{ws.title}:{row_num}] o'qishda xato: {exc}")
                    continue
                if entry is None:
                    continue
                if entry["key"] in by_key:
                    self.stats["duplicates"] += 1
                by_key[entry["key"]] = entry

        return list(by_key.values())

    def _report_duplicate_leads(self):
        """Bir telefonda bir nechta Lead qolganini xabar qiladi.

        Telefonlar to'g'rilangach, ilgari har xil ko'ringan yozuvlar bir xil raqamga
        tushib qolishi mumkin. Avtomatik birlashtirilmaydi — qo'lda ko'rib chiqiladi.
        """
        dupes = (
            Lead.objects.exclude(phone=NO_PHONE).values("phone").annotate(n=Count("id")).filter(n__gt=1).order_by("-n")
        )
        rows = list(dupes[:20])
        if not rows:
            return

        total = dupes.count()
        self.stdout.write(
            self.style.WARNING(f"\n{total} ta telefon bazada bir nechta leadga tegishli — qo'lda birlashtirish kerak:")
        )
        for row in rows:
            names = list(Lead.objects.filter(phone=row["phone"]).values_list("id", "full_name")[:4])
            self.stdout.write(f"   {row['phone']} — {row['n']} ta: {names}")
        if total > len(rows):
            self.stdout.write(f"   ... yana {total - len(rows)} ta")

    def _report_by_sheet(self, entries):
        per_sheet = {}
        for entry in entries:
            bucket = per_sheet.setdefault(entry["sheet"], {"created": 0, "updated": 0, "unchanged": 0})
            if entry.get("outcome"):
                bucket[entry["outcome"]] += 1
        for title, total in self.sheet_rows.items():
            b = per_sheet.get(title, {"created": 0, "updated": 0, "unchanged": 0})
            self.stdout.write(
                f"[{title}] {total} qator -> {b['created']} yangi, {b['updated']} yangilandi, "
                f"{b['unchanged']} o'zgarishsiz"
            )

    # --- qator ---------------------------------------------------------

    def _parse_row(self, data, sheet_title, row_num):
        full_name = normalize(data.get("full_name"))
        org_name = normalize(data.get("org_name"))
        phone = clean_phone(data.get("phone"))
        legacy = legacy_phone(data.get("phone"))

        # Bo'lim sarlavhalari va bo'sh qatorlar: na ism, na tashkilot, na telefon
        if not full_name and not org_name and phone == NO_PHONE:
            self.stats["skipped"] += 1
            return

        display_name = full_name or org_name or NO_PHONE
        status, sub_status, status_note = parse_status(data.get("status"))
        source = SOURCE_MAP.get(norm(data.get("source")), "Boshqa")
        subsidiya = parse_subsidiya(data.get("subsidiya"))

        manager_raw = normalize(data.get("manager"))
        owner = None
        if manager_raw and norm(manager_raw) not in FAKE_MANAGERS:
            owner = self.users.get(norm(manager_raw))
            if not owner:
                self.unmatched[manager_raw] = self.unmatched.get(manager_raw, 0) + 1

        contact_dt = make_aware(data.get("contact_date"))

        meeting_col = data.get("meeting")
        meeting_at = make_aware(meeting_col) if isinstance(meeting_col, datetime) else None
        meeting_note = None if meeting_at else (normalize(meeting_col) or None)

        structured_comments, general_note = parse_comments(data.get("note"))

        # Tashkilot/mahalla nomi ism ustunidan alohida bo'lsa, izohda saqlanadi
        note_parts = [p for p in (org_name if full_name else None, status_note, general_note) if p]
        note = "\n".join(note_parts) or None

        values = {
            "full_name": display_name,
            "source": source,
            "subsidiya": subsidiya,
            "status": status,
            "sub_status": sub_status,
            "owner": owner,
            "assignee": owner,
            "note": note,
            "meeting_at": meeting_at,
            "contacted_at": contact_dt,
        }

        return {
            "key": phone if phone != NO_PHONE else f"name:{norm(display_name)}",
            "sheet": sheet_title,
            "row_num": row_num,
            "phone": phone,
            "legacy_phone": legacy,
            "display_name": display_name,
            "values": values,
            "owner": owner,
            "contact_dt": contact_dt,
            "meeting_note": meeting_note,
            "comments": structured_comments,
        }

    def _apply_row(self, entry):
        lead = self._find_lead(entry["phone"], entry["legacy_phone"], entry["display_name"])
        if lead:
            entry["outcome"] = self._update_lead(lead, entry["values"], entry["sheet"], entry["row_num"])
        else:
            lead = self._create_lead(entry)
            entry["outcome"] = "created"

        self._sync_comments(lead, entry["meeting_note"], entry["comments"], entry["contact_dt"], entry["owner"])

    def _find_lead(self, phone, legacy, display_name):
        if phone == NO_PHONE:
            # Telefonsiz qatorlar takror importda dublikat bo'lmasligi uchun ism bo'yicha qidiriladi
            return Lead.objects.filter(phone=NO_PHONE, full_name__iexact=display_name).order_by("id").first()

        lead = Lead.objects.filter(phone=phone).order_by("id").first()
        if lead or legacy == phone:
            return lead

        # Prodda eski (ortiqcha nolli) formatda saqlangan bo'lsa — topib, raqamini to'g'rilaymiz
        lead = Lead.objects.filter(phone=legacy).order_by("id").first()
        if lead:
            lead.phone = phone
            lead.save(update_fields=["phone"])
            self.stats["phones_fixed"] += 1
        return lead

    def _create_lead(self, entry):
        phone, values = entry["phone"], entry["values"]
        contact_dt, owner = entry["contact_dt"], entry["owner"]
        sheet_title, row_num = entry["sheet"], entry["row_num"]

        lead = Lead.objects.create(phone=phone, board=Lead.BOARD_SALES, score=0, **values)
        if contact_dt:
            Lead.objects.filter(pk=lead.pk).update(created_at=contact_dt)
        self.stats["created"] += 1

        if self.write_events:
            ev = LeadEvent.objects.create(lead=lead, type=LeadEvent.TYPE_CREATED, by=owner)
            self.stats["events"] += 1
            if contact_dt:
                LeadEvent.objects.filter(pk=ev.pk).update(at=contact_dt)
            if values["meeting_at"]:
                LeadEvent.objects.create(
                    lead=lead,
                    type=LeadEvent.TYPE_MEETING,
                    meeting_at=values["meeting_at"],
                    meeting_type="Ofisda",
                    by=owner,
                )
                self.stats["events"] += 1

        self.stdout.write(
            f"  [{sheet_title}:{row_num}] yangi: {values['full_name']} | {phone} | "
            f"{values['status']}/{values['sub_status']}"
        )
        return lead

    def _update_lead(self, lead, values, sheet_title, row_num):
        old = {"status": lead.status, "sub_status": lead.sub_status}
        changed = []

        for field in UPDATABLE_FIELDS:
            new = values.get(field)
            # Exceldagi bo'sh katak proddagi qiymatni o'chirmaydi
            if new in (None, "") or (field == "subsidiya" and not new):
                continue
            current = getattr(lead, field)
            if current == new:
                continue
            if not self.excel_wins and current not in (None, "", False):
                continue
            setattr(lead, field, new)
            changed.append(field)

        if not changed:
            self.stats["unchanged"] += 1
            return "unchanged"

        lead.save(update_fields=changed)
        self.stats["updated"] += 1

        if self.write_events:
            for field, ev_type in (("status", LeadEvent.TYPE_STATUS), ("sub_status", LeadEvent.TYPE_SUB_STATUS)):
                if field in changed:
                    LeadEvent.objects.create(
                        lead=lead,
                        type=ev_type,
                        from_value=old[field],
                        to_value=getattr(lead, field),
                        by=lead.owner,
                    )
                    self.stats["events"] += 1
            if "meeting_at" in changed:
                LeadEvent.objects.create(
                    lead=lead,
                    type=LeadEvent.TYPE_MEETING,
                    meeting_at=lead.meeting_at,
                    meeting_type="Ofisda",
                    by=lead.owner,
                )
                self.stats["events"] += 1

        self.stdout.write(f"  [{sheet_title}:{row_num}] yangilandi: {lead.full_name} -> {', '.join(changed)}")
        return "updated"

    def _sync_comments(self, lead, meeting_note, structured_comments, contact_dt, owner):
        """Izohlarni event sifatida qo'shadi; takroriy importda dublikat yaratmaydi."""
        if not self.write_events:
            return

        existing = set(LeadEvent.objects.filter(lead=lead, type=LeadEvent.TYPE_COMMENT).values_list("text", flat=True))

        pending = []
        if meeting_note:
            pending.append((meeting_note, owner, contact_dt))
        for date_str, time_str, commentor_name, text in structured_comments:
            try:
                at = timezone.make_aware(datetime.strptime(f"{date_str} {time_str}", "%d.%m.%Y %H:%M"))
            except ValueError:
                at = contact_dt or timezone.now()
            pending.append((text.strip(), self.users.get(norm(commentor_name)) or owner, at))

        for text, by, at in pending:
            if not text or text in existing:
                continue
            existing.add(text)
            ev = LeadEvent.objects.create(lead=lead, type=LeadEvent.TYPE_COMMENT, text=text, by=by)
            self.stats["events"] += 1
            if at:
                LeadEvent.objects.filter(pk=ev.pk).update(at=at)
