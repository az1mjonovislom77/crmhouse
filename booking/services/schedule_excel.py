from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from booking.services.schedule import booking_schedule, down_payment_amount

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
BOLD = Font(bold=True)


def _fmt_plain(value, decimals=0):
    q = Decimal(10) ** -decimals
    v = Decimal(value or 0).quantize(q, rounding=ROUND_HALF_UP)
    sign = "-" if v < 0 else ""
    v = abs(v)
    int_part, _, dec_part = f"{v:.{decimals}f}".partition(".")
    int_part = f"{int(int_part):,}"
    if decimals:
        return f"{sign}{int_part},{dec_part}"
    return f"{sign}{int_part}"


def _fmt_som(value):
    v = Decimal(value or 0).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    sign = "-" if v < 0 else ""
    v = abs(v)
    int_part, _, dec_part = f"{v:.2f}".partition(".")
    groups = []
    s = int_part
    while len(s) > 3:
        groups.insert(0, s[-3:])
        s = s[:-3]
    groups.insert(0, s)
    grouped = " ".join(groups)
    return f"{sign}{grouped},{dec_part} so'm"


def _cell(ws, coord, value, *, bold=False, align=CENTER):
    cell = ws[coord]
    cell.value = value
    cell.border = BORDER
    cell.alignment = align
    if bold:
        cell.font = BOLD
    return cell


def build_schedule_xlsx(booking, today=None):
    today = today or timezone.localdate()
    data = booking_schedule(booking, today=today)
    home = booking.home

    wb = Workbook()
    ws = wb.active
    ws.title = "To'lov jadvali"

    widths = [6, 16, 20, 16, 22, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.merge_cells("A1:F1")
    _cell(ws, "A1", "TO'LOV JADVALI", bold=True)
    ws["A1"].font = Font(bold=True, size=14)
    ws.row_dimensions[1].height = 24

    _cell(ws, "A2", "Uy raqami", bold=True, align=LEFT)
    ws.merge_cells("B2:B2")
    _cell(ws, "B2", home.home_number)
    _cell(ws, "C2", "KV-M", bold=True)
    _cell(ws, "D2", "1 KV-M narxi", bold=True)
    ws.merge_cells("E2:F2")
    _cell(ws, "E2", "Sana", bold=True)

    _cell(ws, "A3", "Uy umumiy narxi", bold=True, align=LEFT)
    _cell(ws, "B3", _fmt_plain(booking.total_price))
    ws.merge_cells("C3:C5")
    _cell(ws, "C3", _fmt_plain(home.area, decimals=2))
    ws.merge_cells("D3:D5")
    _cell(ws, "D3", _fmt_plain(home.price_per_sqm or booking.price_per_m2))
    ws.merge_cells("E3:F5")
    _cell(ws, "E3", today.strftime("%d.%m.%Y"))

    down_payment = down_payment_amount(booking)
    _cell(ws, "A4", "Boshlang'ich to'lov", bold=True, align=LEFT)
    _cell(ws, "B4", _fmt_plain(down_payment))

    remaining = data["kpi"]["total_planned"] - down_payment
    _cell(ws, "A5", "Qolgan summa", bold=True, align=LEFT)
    _cell(ws, "B5", _fmt_plain(remaining))

    header_row = 7
    headers = ["№", "To'lov sanasi", "To'lov", "Foizlar", "Qolgan summa", "Imzo"]
    for idx, title in enumerate(headers, start=1):
        _cell(ws, f"{get_column_letter(idx)}{header_row}", title, bold=True)

    row = header_row + 1
    for installment in data["installments"]:
        _cell(ws, f"A{row}", installment["no"])
        _cell(ws, f"B{row}", installment["due_date"].strftime("%d.%m.%Y"))
        _cell(ws, f"C{row}", _fmt_som(installment["planned_amount"]))
        _cell(ws, f"D{row}", _fmt_som(0))
        _cell(ws, f"E{row}", _fmt_som(installment["balance_after"]))
        _cell(ws, f"F{row}", "")
        row += 1

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
