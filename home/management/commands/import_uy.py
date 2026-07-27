import os
import re

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from home.models import Home
from projects.models.project_models import Block, Floors


class Command(BaseCommand):
    help = "uy.xlsx fayldan Home modeliga ma'lumot yuklash"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="uy.xlsx",
            help="Excel fayl nomi yoki to'liq yo'l (default: uy.xlsx)",
        )
        parser.add_argument(
            "--block-prefix",
            default="Block - ",
            help='Block title prefiks (masalan: "Block - " -> "Block - A"). Default: "Block - "',
        )
        parser.add_argument(
            "--project-id",
            type=int,
            default=3,
            help="Block qidirishda cheklash uchun Project ID. Default: 3",
        )
        parser.add_argument(
            "--organization",
            default="Qamashi Xonadonlar",
            help='Block qidirishda faqat shu organizationga tegishli bloklar. Default: "Qamashi Xonadonlar"',
        )
        parser.add_argument(
            "--total-price",
            action="store_true",
            default=False,
            help="price ustuni umumiy narx bo'lsa, price_per_sqm = price / area hisoblanadi",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            default=False,
            help="Bazaga hech narsa yozmaydi, faqat nima bo'lishini hisoblab ko'rsatadi",
        )

    def handle(self, *args, **options):
        file_path = options["file"]
        if not os.path.isabs(file_path):
            file_path = os.path.join(settings.BASE_DIR, file_path)

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"Fayl topilmadi: {file_path}"))
            return

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        self.stdout.write(f"Ustunlar: {headers}")
        self.stdout.write(f"Jami qatorlar: {ws.max_row - 1}")

        block_prefix = options["block_prefix"]
        project_id = options["project_id"]
        organization_name = options["organization"]
        use_total_price = options["total_price"]
        dry_run = options["dry_run"]

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN rejimi: bazaga hech narsa yozilmaydi"))

        created_count = 0
        updated_count = 0
        skipped_count = 0
        error_count = 0

        for row_idx in range(2, ws.max_row + 1):
            row = {headers[c]: ws.cell(row_idx, c + 1).value for c in range(len(headers))}

            if not row.get("home_number"):
                self.stdout.write(
                    self.style.WARNING(f"Qator {row_idx}: 'home_number' bo'sh, o'tkazib yuborildi | ma'lumot: {row}")
                )
                skipped_count += 1
                continue

            try:
                floor_number = int(row["floor"])
                home_number = int(row["home_number"])
                area = float(row["area"])
                base_entrance = int(row["entrance"])
                department = int(row.get("department") or 1)
                entrance = base_entrance + (department - 1) * 2

                raw_price = row.get("price") or 0
                if isinstance(raw_price, str):
                    raw_price = raw_price.replace(" ", "").replace("\xa0", "").replace(",", ".")
                price = float(raw_price)
                price_per_sqm = round(price / area, 2) if use_total_price and area else price

                room_str = str(row.get("room") or "1")
                match = re.search(r"\d+", room_str)
                rooms = int(match.group()) if match else 1

                if dry_run:
                    floor_obj = Floors.objects.filter(number=floor_number).first()
                else:
                    floor_obj, _ = Floors.objects.get_or_create(number=floor_number)

                block_letter = str(row.get("block") or "").strip()
                block_title = f"{block_prefix}{block_letter}"
                block_qs = Block.objects.filter(title=block_title)
                if organization_name:
                    block_qs = block_qs.filter(projects__organization__name=organization_name)
                if project_id:
                    block_qs = block_qs.filter(projects_id=project_id)
                block_obj = block_qs.first()

                if not block_obj:
                    self.stdout.write(
                        self.style.WARNING(
                            f"Qator {row_idx}: Block '{block_title}' "
                            f"('{organization_name}' organizationida) topilmadi, o'tkazib yuborildi"
                        )
                    )
                    skipped_count += 1
                    continue

                if dry_run:
                    exists = Home.objects.filter(
                        home_number=home_number,
                        floor=floor_obj,
                        blocks=block_obj,
                        entrance=entrance,
                    ).exists()
                    action = "YANGILANADI" if exists else "YARATILADI"
                    self.stdout.write(
                        f"Qator {row_idx}: {action} | block='{block_obj.title}' (id={block_obj.id}) "
                        f"floor={floor_number} home_number={home_number} "
                        f"area={area} rooms={rooms} price_per_sqm={price_per_sqm} "
                        f"department={department} entrance={entrance}"
                    )
                    if exists:
                        updated_count += 1
                    else:
                        created_count += 1
                    continue

                _, created = Home.objects.update_or_create(
                    home_number=home_number,
                    floor=floor_obj,
                    blocks=block_obj,
                    entrance=entrance,
                    defaults={
                        "area": area,
                        "rooms": rooms,
                        "price_per_sqm": price_per_sqm,
                    },
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

            except (ValueError, TypeError, KeyError) as e:
                self.stdout.write(self.style.ERROR(f"Qator {row_idx} xato: {e} | ma'lumot: {row}"))
                error_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"\nImport tugadi!\n"
                f"  Yaratildi:          {created_count}\n"
                f"  Yangilandi:         {updated_count}\n"
                f"  O'tkazib yuborildi: {skipped_count}\n"
                f"  Xato:               {error_count}"
            )
        )
